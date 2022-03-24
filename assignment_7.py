import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl

s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
driver = webdriver.Chrome(service=s)

driver.maximize_window()
path = "/Users/adamroberts/Documents/Selenium/assignment_sheet.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
nrows = sheet_obj.max_row

for i in range(2, nrows + 1):
    action_type = sheet_obj.cell(row=i, column=2).value
    data = sheet_obj.cell(row=i, column=3).value
    xpath = sheet_obj.cell(row=i, column=4).value
    if action_type == "Open URL":
        try:
            driver.get(data)
            sheet_obj.cell(row=i, column=5).value = "Pass"
        except Exception as e:
            sheet_obj.cell(row=i, column=5).value = "Fail"
            sheet_obj.cell(row=i, column=6).value = e.args[0]
    if action_type == "TEXT":
        try:
            element = driver.find_element(By.XPATH, xpath)
            element.send_keys(data)
            sheet_obj.cell(row=i, column=5).value = "Pass"
        except Exception as e:
            sheet_obj.cell(row=i, column=5).value = "Fail"
            sheet_obj.cell(row=i, column=6).value = e.args[0]
                # sheet_obj.cell(row=i, column=5).font = Font(name="Tahoma", size=12, color="00339966")
                # sheet_obj.cell(row=i, column=5).style = "Hyperlink"

    if action_type == "CLICK":
        try:
            element = driver.find_element(By.XPATH, xpath)
            actions = ActionChains(driver)
            actions.move_to_element(element)
            actions.click()
            actions.perform()
            driver.find_element(By.XPATH, xpath).click()
            sheet_obj.cell(row=i, column=5).value = "Pass"
        except Exception as e:
            print(e.args[0])
            sheet_obj.cell(row=i, column=5).value = "Fail"
            sheet_obj.cell(row=i, column=6).value = e.args[0]


wb_obj.save("/Users/adamroberts/Documents/Selenium/assignment_sheet.xlsx")
time.sleep(5)
driver.quit()